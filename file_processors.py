import mimetypes
import os
import logging
from typing import Dict, Tuple, Optional, Any
from fastapi.concurrency import run_in_threadpool
from fastapi import Request
import io
import uuid
import subprocess

# 追加ログ用ロガー
logger = logging.getLogger(__name__)

import fitz  # PyMuPDF
from docx import Document as DocxDocument
from openpyxl import load_workbook
import boto3 # AWS SDK
from botocore.exceptions import ClientError, BotoCoreError # AWS SDK exceptions

# --- 定数定義 ---
MAX_TEXT_FILE_SIZE_MB = 10  # TXT, MD, code files
MAX_PDF_SIZE_MB = 50       # PDF files for text extraction
MAX_DOCX_SIZE_MB = 20      # DOCX files
MAX_XLSX_SIZE_MB = 20      # XLSX files
MAX_IMAGE_SIZE_MB = 20     # Images for Gemini Vision
MAX_AUDIO_SIZE_MB = 25     # Audio for Whisper

MB_TO_BYTES = 1024 * 1024

SUPPORTED_TEXT_EXT = ['.txt', '.md', '.py', '.js', '.html', '.css', '.json', '.csv', '.log', '.rtf']
SUPPORTED_PDF_EXT = ['.pdf']
SUPPORTED_DOCX_EXT = ['.docx']
SUPPORTED_XLSX_EXT = ['.xlsx']
SUPPORTED_IMAGE_MIMETYPES = ['image/png', 'image/jpeg', 'image/webp', 'image/gif', 'image/bmp']
SUPPORTED_AUDIO_MIMETYPES = ['audio/mpeg', 'audio/wav', 'audio/mp4', 'audio/x-m4a', 'audio/ogg', 'audio/flac', 'audio/webm']

MIN_TEXT_LENGTH_FOR_RICH_EXTRACTION = 50
TEXTRACT_MAX_FILE_SIZE_BYTES = 5 * 1024 * 1024 # 5MB for AWS Textract synchronous API

# --- ヘルパー関数 ---
def get_file_details(filename: str, content: bytes) -> Tuple[str, str, Optional[str], int]:
    """ファイル名、拡張子、MIMEタイプ、サイズを取得"""
    name_no_ext, extension = os.path.splitext(filename)
    extension = extension.lower()
    mime_type, _ = mimetypes.guess_type(filename)
    size_bytes = len(content)
    return name_no_ext, extension, mime_type, size_bytes


def create_error_response(message: str, status_code: int = 400, ai_used: str = "N/A") -> Dict[str, Any]:
    logger.debug("Error response: %s (AI: %s, status=%s)", message, ai_used, status_code)
    return {
        "processed_content": None,
        "original_filename": None,
        "content_type": None,
        "size_bytes": None,
        "processing_ai": ai_used,
        "error": message,
        "status_code": status_code,
    }


def create_success_response(processed_content: str, ai_used: str, original_filename: str, content_type: Optional[str], size_bytes: int) -> Dict[str, Any]:
    logger.debug(
        "Success response (%s) len=%d bytes via %s",
        original_filename,
        len(processed_content) if processed_content else 0,
        ai_used,
    )
    return {
        "processed_content": processed_content,
        "original_filename": original_filename,
        "content_type": content_type,
        "size_bytes": size_bytes,
        "processing_ai": ai_used,
        "error": None,
        "status_code": 200,
    }


def _log_text_preview(text: Optional[str], method: str) -> None:
    """Extract length and first 200 chars for debugging."""
    if text is None:
        logger.debug("%s produced no text", method)
    else:
        preview = text.replace("\n", " ")[:200]
        logger.debug("%s produced %d chars: %s%s", method, len(text), preview, "..." if len(text) > 200 else "")


# --- 各ファイルタイプ処理関数 ---
async def process_text_file(filename: str, content: bytes, extension: str, size_bytes: int) -> Dict[str, Any]:
    if size_bytes > MAX_TEXT_FILE_SIZE_MB * MB_TO_BYTES:
        return create_error_response(
            f"テキストファイル ({extension}) は {MAX_TEXT_FILE_SIZE_MB}MB までしかアップロードできません。", 413
        )
    try:
        text_content = ""
        try:
            text_content = content.decode('utf-8')
        except UnicodeDecodeError:
            try:
                text_content = content.decode('shift-jis')
            except UnicodeDecodeError:
                text_content = content.decode('latin-1', errors='replace')

        return create_success_response(text_content, "Direct Read", filename, mimetypes.guess_type(filename)[0], size_bytes)
    except Exception as e:  # pragma: no cover - unexpected decode issues
        return create_error_response(f"テキストファイル処理エラー: {str(e)}", 500)


async def _process_pdf_with_fitz(content: bytes, filename: str) -> Tuple[Optional[str], str]:
    logger.info("Starting PyMuPDF extraction for '%s'", filename)
    try:
        def extract_text_from_pdf_fitz_sync():
            doc = fitz.open("pdf", content)  # type: ignore
            text_parts = []
            for page_num in range(len(doc)):
                page = doc.load_page(page_num)
                page_text = page.get_text("text", sort=True)
                text_parts.append(page_text)
                logger.debug(
                    "PyMuPDF page %d extracted %d characters",
                    page_num + 1,
                    len(page_text),
                )
            doc.close()
            return "\n--- Page Break ---\n".join(text_parts)

        extracted_text = await run_in_threadpool(extract_text_from_pdf_fitz_sync)
        if extracted_text and extracted_text.strip():
            logger.info("PyMuPDF finished for '%s'. Extracted %d characters.", filename, len(extracted_text))
            _log_text_preview(extracted_text, "PyMuPDF")
        elif extracted_text is not None: # Extracted, but empty or only whitespace
            logger.info("PyMuPDF extraction for '%s' yielded no text.", filename)
            _log_text_preview(extracted_text, "PyMuPDF") # Log it anyway for preview
        else: # None was returned, implies an error before text assignment
            logger.info("PyMuPDF extraction for '%s' returned no text object (None).", filename)
            _log_text_preview(None, "PyMuPDF")
        return extracted_text, "PyMuPDF (fitz)"
    except Exception as e:
        logger.error("PyMuPDF (fitz) error during PDF processing for '%s': %s", filename, e, exc_info=True)
        return None, "PyMuPDF (fitz) - Error"


async def _process_pdf_with_pandoc(content: bytes, filename: str) -> Tuple[Optional[str], str]:
    temp_pdf_path = None
    logger.info("Starting Pandoc extraction for '%s'", filename)
    try:
        temp_dir = "temp_pandoc_files"
        os.makedirs(temp_dir, exist_ok=True)
        temp_pdf_filename = f"{uuid.uuid4()}.pdf"
        temp_pdf_path = os.path.join(temp_dir, temp_pdf_filename)
        with open(temp_pdf_path, "wb") as f:
            f.write(content)

        cmd = ["pandoc", temp_pdf_path, "-t", "plain", "--wrap=none"]

        def run_pandoc_command_sync():
            logger.debug("Running command: %s", " ".join(cmd))
            process = subprocess.run(cmd, capture_output=True, text=True, check=False, timeout=60)
            if process.returncode != 0:
                logger.warning(
                    "Pandoc command failed for '%s' (code %s). Stderr: %s",
                    filename,
                    process.returncode,
                    process.stderr.strip(),
                )
                return None
            return process.stdout

        extracted_text = await run_in_threadpool(run_pandoc_command_sync)
        if extracted_text and extracted_text.strip():
            logger.info("Pandoc finished for '%s'. Extracted %d characters.", filename, len(extracted_text))
            _log_text_preview(extracted_text, "Pandoc")
        elif extracted_text is not None: # Extracted, but empty or only whitespace
            logger.info("Pandoc extraction for '%s' yielded no text.", filename)
            _log_text_preview(extracted_text, "Pandoc")
        else: # None was returned
            logger.info("Pandoc extraction for '%s' returned no text object (None).", filename)
            _log_text_preview(None, "Pandoc")
        return extracted_text, "Pandoc"
    except subprocess.TimeoutExpired:
        logger.error("Pandoc processing for '%s' timed out.", filename, exc_info=True)
        return None, "Pandoc - Timeout"
    except FileNotFoundError:
        logger.error("Pandoc command not found. Ensure it's installed and in PATH.", exc_info=True)
        return None, "Pandoc - Not Found"
    except Exception as e:
        logger.error("Pandoc processing error for '%s': %s", filename, e, exc_info=True)
        return None, "Pandoc - Error"
    finally:
        if temp_pdf_path and os.path.exists(temp_pdf_path):
            try:
                os.remove(temp_pdf_path)
            except OSError as e_remove:
                logger.warning("Error removing temp pandoc file %s: %s", temp_pdf_path, e_remove)


async def _process_pdf_with_textract(filename: str, content: bytes, textract_client: Optional[boto3.client]) -> Tuple[Optional[str], str]:
    if not textract_client:
        logger.warning("AWS Textract client not configured. Cannot process '%s'.", filename)
        return None, "AWS Textract (Not Configured)"

    logger.info("Starting AWS Textract processing for '%s'", filename)
    try:
        # analyze_document is synchronous, run in threadpool
        def call_textract_sync():
            return textract_client.analyze_document(
                Document={'Bytes': content},
                FeatureTypes=['TEXT']
            )

        response = await run_in_threadpool(call_textract_sync)

        extracted_lines = []
        if 'Blocks' in response:
            for block in response['Blocks']:
                if block['BlockType'] == 'LINE':
                    if 'Text' in block and block['Text']:
                        extracted_lines.append(block['Text'])
        
        extracted_text = "\n".join(extracted_lines)

        if extracted_text and extracted_text.strip():
            logger.info("AWS Textract finished for '%s'. Extracted %d characters.", filename, len(extracted_text))
            _log_text_preview(extracted_text, "AWS Textract")
            return extracted_text, "AWS Textract"
        elif extracted_text is not None: # Extracted, but empty or only whitespace
            logger.info("AWS Textract extraction for '%s' yielded no text.", filename)
            _log_text_preview(extracted_text, "AWS Textract") # Log it anyway for preview
            return extracted_text, "AWS Textract" # Return empty string rather than None
        else: # Should not happen if logic is correct, means no 'LINE' blocks with text
            logger.info("AWS Textract extraction for '%s' resulted in no text blocks.", filename)
            _log_text_preview(None, "AWS Textract")
            return None, "AWS Textract - No Text Blocks"

    except (ClientError, BotoCoreError) as e: # Catch specific AWS SDK errors
        error_message = str(e)
        logger.error("AWS Textract API error for '%s': %s", filename, error_message, exc_info=True)
        # Check for common error types to provide more specific feedback
        if "UnsupportedDocumentException" in error_message:
            return None, "AWS Textract - Unsupported Document"
        elif "DocumentTooLargeException" in error_message:
            return None, "AWS Textract - Document Too Large"
        elif "BadDocumentException" in error_message:
            return None, "AWS Textract - Bad Document"
        return None, f"AWS Textract - API Error ({e.__class__.__name__})"
    except Exception as e:
        logger.error("Unexpected error during AWS Textract processing for '%s': %s", filename, e, exc_info=True)
        return None, f"AWS Textract - Error ({e.__class__.__name__})"


async def process_pdf_file(filename: str, content: bytes, size_bytes: int, textract_client: Optional[Any]) -> Dict[str, Any]:
    logger.info("Processing PDF file '%s' (%d bytes)", filename, size_bytes)
    if size_bytes > MAX_PDF_SIZE_MB * MB_TO_BYTES:
        return create_error_response(
            f"PDFファイルは {MAX_PDF_SIZE_MB}MB までしかアップロードできません。", 413, "N/A"
        )

    current_best_text: Optional[str] = None
    method_used: str = "N/A"
    tried_methods: list[str] = []

    # 1. Try PyMuPDF (fitz)
    fitz_text, fitz_method = await _process_pdf_with_fitz(content, filename)
    tried_methods.append(fitz_method)
    if fitz_text and fitz_text.strip():
        current_best_text = fitz_text
        method_used = fitz_method
        if len(current_best_text.strip()) >= MIN_TEXT_LENGTH_FOR_RICH_EXTRACTION:
            logger.info("PyMuPDF successfully extracted sufficient text for '%s'.", filename)
            _log_text_preview(current_best_text, method_used)
            return create_success_response(current_best_text, method_used, filename, "application/pdf", size_bytes)

    logger.info(
        "PyMuPDF for '%s' yielded %s. Current best text length: %d. (Method: %s)",
        filename,
        "no text" if not fitz_text or not fitz_text.strip() else "short text",
        len(current_best_text.strip()) if current_best_text and current_best_text.strip() else 0,
        fitz_method
    )

    # 2. Try Pandoc if PyMuPDF failed or text is insufficient
    should_try_pandoc = (current_best_text is None or
                         not current_best_text.strip() or
                         len(current_best_text.strip()) < MIN_TEXT_LENGTH_FOR_RICH_EXTRACTION)

    if should_try_pandoc:
        logger.info("Trying Pandoc for '%s'.", filename)
        pandoc_text, pandoc_method = await _process_pdf_with_pandoc(content, filename)
        tried_methods.append(pandoc_method)
        if pandoc_text and pandoc_text.strip():
            # Use Pandoc result if it's better than current_best_text (which might be empty from fitz)
            if not current_best_text or not current_best_text.strip() or len(pandoc_text.strip()) > len(current_best_text.strip()):
                logger.info("Pandoc yielded better or initial text for '%s'.", filename)
                current_best_text = pandoc_text
                method_used = pandoc_method
            else:
                logger.info("Pandoc text for '%s' was not better than PyMuPDF's.", filename)
            
            if current_best_text and current_best_text.strip() and len(current_best_text.strip()) >= MIN_TEXT_LENGTH_FOR_RICH_EXTRACTION:
                logger.info("Pandoc successfully extracted sufficient text for '%s'.", filename)
                _log_text_preview(current_best_text, method_used)
                return create_success_response(current_best_text, method_used, filename, "application/pdf", size_bytes)
        else:
            logger.warning("Pandoc failed or yielded empty text for '%s' (Method: %s).", filename, pandoc_method)
    
    logger.info(
        "After Pandoc for '%s', current best text length: %d. (Method: %s)",
        filename,
        len(current_best_text.strip()) if current_best_text and current_best_text.strip() else 0,
        method_used
    )

    # 3. Try AWS Textract if Pandoc (or PyMuPDF) failed or text is insufficient
    should_try_textract = (current_best_text is None or
                           not current_best_text.strip() or
                           len(current_best_text.strip()) < MIN_TEXT_LENGTH_FOR_RICH_EXTRACTION)

    if should_try_textract:
        if size_bytes > TEXTRACT_MAX_FILE_SIZE_BYTES:
            logger.info(
                "File size (%d bytes) for '%s' exceeds Textract limit (%d bytes). Skipping Textract.",
                size_bytes, filename, TEXTRACT_MAX_FILE_SIZE_BYTES
            )
            # tried_methods.append("AWS Textract (Skipped - Size Limit)") # Optional: add to tried_methods
        else:
            logger.info("Trying AWS Textract for '%s'.", filename)
            textract_text, textract_method = await _process_pdf_with_textract(filename, content, textract_client)
            tried_methods.append(textract_method)
            if textract_text and textract_text.strip():
                 # Use Textract result if it's better
                if not current_best_text or not current_best_text.strip() or len(textract_text.strip()) > len(current_best_text.strip()):
                    logger.info("AWS Textract yielded better or initial text for '%s'.", filename)
                    current_best_text = textract_text
                    method_used = textract_method
                else:
                    logger.info("AWS Textract text for '%s' was not better than previous methods.", filename)

                # No need to check MIN_TEXT_LENGTH_FOR_RICH_EXTRACTION here, this is the last resort.
            else:
                logger.warning("AWS Textract failed or yielded empty text for '%s' (Method: %s).", filename, textract_method)

    # Final check and response
    if current_best_text and current_best_text.strip():
        logger.info(
            "Final extracted text for '%s' using %s has length %d.",
            filename, method_used, len(current_best_text.strip())
        )
        _log_text_preview(current_best_text, method_used)
        return create_success_response(current_best_text, method_used, filename, "application/pdf", size_bytes)
    else:
        final_method_info = " | ".join(tried_methods) if tried_methods else "N/A"
        logger.error("All PDF extraction methods failed for '%s'. Tried: %s", filename, final_method_info)
        return create_error_response(
            f"PDFファイル「{filename}」からテキストを抽出できませんでした。試行したすべての方法で有効なテキストが得られませんでした。",
            422,
            final_method_info + " - All Failed"
        )

async def process_docx_file(filename: str, content: bytes, size_bytes: int) -> Dict[str, Any]:
    if size_bytes > MAX_DOCX_SIZE_MB * MB_TO_BYTES:
        return create_error_response(
            f"Word (docx) ファイルは {MAX_DOCX_SIZE_MB}MB までしかアップロードできません。", 413
        )
    try:
        def extract_text_from_docx():
            doc = DocxDocument(io.BytesIO(content))
            text = "\n".join([para.text for para in doc.paragraphs])
            return text

        extracted_text = await run_in_threadpool(extract_text_from_docx)
        return create_success_response(extracted_text, "python-docx", filename, mimetypes.guess_type(filename)[0], size_bytes)
    except Exception as e:
        return create_error_response(f"Word (docx) ファイル処理エラー: {str(e)}", 500, "python-docx")


async def process_xlsx_file(filename: str, content: bytes, size_bytes: int) -> Dict[str, Any]:
    if size_bytes > MAX_XLSX_SIZE_MB * MB_TO_BYTES:
        return create_error_response(
            f"Excel (xlsx) ファイルは {MAX_XLSX_SIZE_MB}MB までしかアップロードできません。", 413
        )
    try:
        def extract_text_from_xlsx():
            workbook = load_workbook(io.BytesIO(content))
            text_parts = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_parts.append(f"--- Sheet: {sheet_name} ---")
                for row in sheet.iter_rows(values_only=True):
                    row_texts = [str(cell) if cell is not None else "" for cell in row]
                    text_parts.append(", ".join(row_texts))
                text_parts.append("\n")
            return "\n".join(text_parts)

        extracted_text = await run_in_threadpool(extract_text_from_xlsx)
        return create_success_response(extracted_text, "openpyxl", filename, mimetypes.guess_type(filename)[0], size_bytes)
    except Exception as e:
        return create_error_response(f"Excel (xlsx) ファイル処理エラー: {str(e)}", 500, "openpyxl")


async def process_image_file(filename: str, content: bytes, mime_type: Optional[str], size_bytes: int, gemini_vision_client: Optional[Any]) -> Dict[str, Any]:
    if not gemini_vision_client:
        return create_error_response("Gemini Visionクライアントが利用できません。", 503, "Gemini Vision (Not Configured)")
    if not mime_type:
        return create_error_response("画像のMIMEタイプが不明です。", 415, "Gemini Vision")
    if size_bytes > MAX_IMAGE_SIZE_MB * MB_TO_BYTES:
        return create_error_response(
            f"画像ファイルは {MAX_IMAGE_SIZE_MB}MB までしかアップロードできません。", 413, "Gemini Vision"
        )
    try:
        image_part = {"mime_type": mime_type, "data": content}
        prompt = "この画像に何が写っているか詳細に説明してください。画像内のテキストも読み取ってください。"

        async def generate_image_description():
            response = await gemini_vision_client.generate_content_async([prompt, image_part])
            return response.text

        description = await generate_image_description()

        return create_success_response(description, "Gemini Vision", filename, mime_type, size_bytes)
    except Exception as e:
        logger.error("Gemini Vision API error: %s", e)
        return create_error_response(f"画像処理エラー (Gemini Vision): {str(e)}", 500, "Gemini Vision")


async def process_audio_file(filename: str, content: bytes, mime_type: Optional[str], size_bytes: int, openai_client: Optional[Any]) -> Dict[str, Any]:
    if not openai_client:
        return create_error_response("OpenAIクライアントが利用できません。", 503, "OpenAI Whisper (Not Configured)")
    if not mime_type:
        return create_error_response("音声のMIMEタイプが不明です。", 415, "OpenAI Whisper")
    if size_bytes > MAX_AUDIO_SIZE_MB * MB_TO_BYTES:
        return create_error_response(
            f"音声ファイルは {MAX_AUDIO_SIZE_MB}MB までしかアップロードできません。", 413, "OpenAI Whisper"
        )
    try:
        audio_file_like = io.BytesIO(content)

        async def transcribe_audio():
            response = await openai_client.audio.transcriptions.create(
                model="whisper-1",
                file=(filename, audio_file_like, mime_type)
            )
            return response.text

        transcript_text = await transcribe_audio()

        return create_success_response(transcript_text, "OpenAI Whisper", filename, mime_type, size_bytes)
    except Exception as e:
        logger.error("OpenAI Whisper API error: %s", e)
        return create_error_response(f"音声処理エラー (OpenAI Whisper): {str(e)}", 500, "OpenAI Whisper")


# --- メイン処理関数 ---
async def stage0_process_file(request: Request, filename: str, content: bytes) -> Dict[str, Any]:
    name_no_ext, extension, mime_type, size_bytes = get_file_details(filename, content)

    logger.debug(
        "stage0: name=%s ext=%s mime=%s size=%d", name_no_ext, extension, mime_type, size_bytes
    )

    openai_client_instance = request.app.state.openai_client
    gemini_vision_client_instance = request.app.state.gemini_vision_client
    textract_client_instance = request.app.state.textract_client

    if extension in SUPPORTED_TEXT_EXT:
        logger.debug("Dispatching to text processor")
        return await process_text_file(filename, content, extension, size_bytes)
    elif extension in SUPPORTED_PDF_EXT:
        logger.debug("Dispatching to PDF processor")
        return await process_pdf_file(filename, content, size_bytes, textract_client_instance)
    elif extension in SUPPORTED_DOCX_EXT:
        logger.debug("Dispatching to DOCX processor")
        return await process_docx_file(filename, content, size_bytes)
    elif extension in SUPPORTED_XLSX_EXT:
        logger.debug("Dispatching to XLSX processor")
        return await process_xlsx_file(filename, content, size_bytes)

    if mime_type:
        if mime_type in SUPPORTED_IMAGE_MIMETYPES:
            logger.debug("Dispatching to image processor")
            return await process_image_file(filename, content, mime_type, size_bytes, gemini_vision_client_instance)
        elif mime_type in SUPPORTED_AUDIO_MIMETYPES:
            logger.debug("Dispatching to audio processor")
            return await process_audio_file(filename, content, mime_type, size_bytes, openai_client_instance)

    unsupported_message = f"この形式のファイル ({extension if extension else mime_type if mime_type else '不明'}) は扱えません。"
    return create_error_response(unsupported_message, 415)
